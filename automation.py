import openpyxl as xl
from openpyxl.chart import Reference, BarChart

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    sheet['d1'] = 'corrected price'
    sheet['f1'] = 'chart'

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3) #this will give the price column
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save('transactions.xlsx')

process_workbook('transactions.xlsx')
